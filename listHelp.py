# -*- coding: utf-8 -*-
"""
Created on Tue Feb 13 08:27:51 2018
    Python系统可以通过dir()、type()函数和__doc__属性获得指定模块(或对象)的各项名称、属性的类
型，以及函数的内置说明。本程序通过利用上述3个函数提起指定模块的信息，并保存在excel文
件内，以便于以后的查询阅读。
    2018-2-14:
        经过2天多的努力完成了我认可的第一个能用的版本。
        发现问题：1.type()函数的参数应是对象类型，不能直接用模块名字符串代替。此问题
    目前没有办法解决。2.module对象转换为str类型后，字符串中存在不能作为Excel.Sheet的
    名称的特殊字符，通过RE提取module名解决。3.Python序列默认从0开始，而Excel单元格序
    列默认从1开始，在处理时要做转换。
    模块内置函数：
    1.ge_help(module):对传入的模块提取相关的帮助内容，并返回含有帮助的元组。
    2.write_help(module,in_help,filename):将传入的帮助内容元组（in_help）写入Excel
    文件（filename）中的表（module）中。
    3.help_to_excel(module,filename):调用get_help、write_help，将module模块的帮助
    写入filename文件（Excel格式）已module命名的表中。
	2018-3-9，修订，扩展提取帮助范围，不限于module。
	2018-3-10，今天，在使用中发现，python中有部分特殊类型，如‘function’、‘_frozen_importlib.ModuleSpec’
    等，他们的实例在使用str(type(module))方式提取的名称仍然是类型名，而不是包含在模块中的索引项名。？？
@author: ccds_stx
"""

def get_module_name(module):
    """
       #2018-3-9,增加本函数用于处理传递过来的模块名，或类型变量
       #功能：判断传入参数的类型，如果是模块（module）、类型（type）则保留原变量，如果是实例则转换为对应的类型（type）
       #输入：module，传入的模块、类型及实例。
       #返回值：一个元组（列表），含3个成员：0.类型变量，即模块、类型；1.名称，模块名、类型名的字符串；2.传入对象类型，module、type或other。
    """
    #提取传入参数的类型，是“module”、“type”或其他
    import re
    module_type_str = str(type(module))  #将传入参数的类型字符串化
    re_name = re.compile(r'\'.*?\'')     #设置提取模块名（类名）等模式
    re_result = re.search(re_name,module_type_str)
    temp_type = re_result.group()[1:-1]  #将得到的参数类型名称存入变量“temp_type”
    #根据参数类型，处理生成返回结果
    if temp_type == 'module':#传入的是模块
        #提取真实模块名
        re_result = re.search(re_name,str(module))
        module_name = re_result.group()[1:-1]
    elif temp_type == 'type':#传入的是类型
        #提取真实类型名
        re_result = re.search(re_name,str(module))
        module_name = re_result.group()[1:-1]
    else: #传入参数是实例
        #下面处理顺序不能更改
        #1.先生成实例的显示名，可能为“真实类型名->实例名”或“类型.实例名”
        """
         此处比较复杂，当类型为“实例”时，有些实例的类型名称表现形式为在模块（类）内部，具有较清晰的层级结构的命令规则，且没有“__name__”属性，
         如openpyxl模块中的“openpyxl.workbook.workbook.Workbook”。还有一些则只有简单的类型名称，而没有在模块（类）内部的层次体现，此情况，其中一些它的“
         __name__”包含了他在模块（类）内的名称，还有一些连“__name__”属性也没有。
         目前处理方式，先尝试按“简单类型+实例.__name__”处理，报错（AttributeError:没有__name__属性）时，直接采用类型名称字符串作显示名。
        """
        try:
            module_name =   temp_type+'->'+module.__name__  #补充实例名,按“真实类型名->实例名”格式处理
        except AttributeError:   #实例没有“__name__”属性，直接使用实例类型
            re_result = re.search(re_name,str(type(module)))
            module_name = re_result.group()[1:-1]
        #2.判断传入参数是否是实例，如果是实例则转换传入参数为类型 ,必须放在提取真实类型名后面
        if  not hasattr(module,'__bases__'):
            module= type(module)
        #3.将temp_type改为'other'做为返回值
        temp_type = 'other'
    #生成返回元组，返回
    return (module,module_name,temp_type)


def format_typestr(type_string):
    """
    2018-3-1,增加本函数，用于将类型名称字符串进行格式化，去除多余内容，仅留下类型名称。
    注意：实参必须时字符串，本函数内不进行类型校验。
    """
    import re
    type_re = re.compile(r'\'.*\'')
    type_match = type_re.search(type_string)
    type_name = type_match.group()[1:-1] #使用字符串切片方式 将两端的“'”
    return type_name

def get_help(module):
    """
    本函数通过dir()函数获得给定的模块(类)的内部成员信息，包括模块（类）内部成员名称、类型、内置帮助信息等内容，并
    保存在列表（或元组）中。
    输入：module，要提取帮助信息的模块（类）。注：object类型。
    输出：返回包含帮助信息的列表（元组）名。
    """
    name_list = dir(module)
    header=('名称','类型','内部帮助')  #定义表头信息
    content = [] #列表，保存获得的所有内容
    content.append(header)
    for i in range(0,len(name_list)):
        x = name_list[i] # 变量x保存名称
        try:  #针对 部分属性处理错误，目前无好的处理办法，直接将属性类型赋值为''
            y = str(type(getattr(module,x,''))) #变量y保存类型,如果属性不存在则直接赋空字符串。
            #2018-3-1 增加格式化类型名称处理
            y = format_typestr(y)
        except:
            y = ''
        #获取帮助的具体内容
        z = getattr(module,x,'')  #获取属性，如果属性不存在则直接赋值为空字符串
        if z != '':
            z = getattr(module,x).__doc__ #变量z保存帮助内容
        content.append((x,y,z))
    return tuple(content)

def write_help(module_name,in_help,filename):
    """
    本函数将获得模块内容帮助内容写入指定Excel文件中。
    输入：
    1.module_name：模块名称，str类型。
    2.in_help：获得的帮助内容，tuple类型。
    3.filename：保存文件名称，即Excel文件名称，str类型。
    输出：字符串，'Ok':写入成功；其它:写入不成功，字符串内容未不成功原因。
    """
    import openpyxl
    from os.path import exists as fileexists
    if not fileexists(filename):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(filename)
    if  module_name in wb.sheetnames:
        return 'Error:已经有相关模块的帮助了，没必要重新写入。'
    else:
        ws = wb.create_sheet(title=module_name)
        for row in range(1,len(in_help)+1):
            for col in range(1,len(in_help[row-1])+1):
                ws.cell(row,col).value = in_help[row-1][col-1]
        wb.save(filename)
        return 'Ok'

def help_to_excel(module,filename):
    """
    完成指定模块的帮助信息的收集以及写入Excel文件的过程。
    输入：(1)module，指定的模块对象。(2)filename，写入帮助的Excel文件名。
    输出：str，“Ok”正常完成处理结束；其他，错误。
    """
    if isinstance(module,(int,float,str,bool,list,tuple,dict,set)):#简单类型：int,float,str,bool,list,tuple,dict,set
        return "简单类型：{}，没必要提取帮助！".format(str(type(module)))
    #调用get_module_name函数，生产参数信息    
    module_info = get_module_name(module)    
    lines = get_help(module_info[0])
    str_write = write_help(module_info[1],lines,filename)
    return str_write


